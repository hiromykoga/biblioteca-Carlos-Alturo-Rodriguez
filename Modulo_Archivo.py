#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
from Modulo_libro import Libro

class Archivo():
	def __init__(self,ruta):
		self.ruta = ruta
		self.columnas = ['Titulo','Autores','Editorial','Codigo_isbn','Cantidad','Color_cinta','Color_rotulo','Clave_coleccion','Clave_autor']
	
	def __str__(self):
		archivo = "ruta: %s"%(self.ruta)
		return archivo
	
	def cargarArchivo(self):
		"""
		Metodo que se utiliza para cargar el archivo
		"""
		try:
			archivoExcel = openpyxl.load_workbook(self.ruta) #carga el archivo de excel
			nombreHojas = archivoExcel.get_sheet_names() #obtiene el nombre de todas las hojas del excel
			hoja = archivoExcel.get_sheet_by_name(nombreHojas[0]) #seleccionamos la primer hoja que es la vamos a utilizar
			print("cargo bn el archivo")
		except :
			print("Error en el momento de cargar el archivo")
			pass

		return hoja

	def validaEstructuraArchivo(self,hoja):
		"""
		Metodo que valida la estructura de la hoja, que tenga la cantidad de columnas correspondientes
		y el nombre de ellas sea la correspondientes
		"""
		primerFila = 1
		columnasCorrectas = True
		cantidad_columnas = hoja.max_column #cuenta la cantidad de columnas del excel
		if cantidad_columnas == 9:
			for columna in range(1,(cantidad_columnas + 1)):
				valor = hoja.cell(row = primerFila, column = columna).value #valor que tiene la celda
				if valor not in self.columnas:
					columnasCorrectas = False
					break
		return columnasCorrectas
				
	def convertirExcelEnDiccionario(self,hoja):
		"""
		Este metodo realiza la validacion de cada valor de cado campo, es decir que
		cumplan con el tipo correcto, ademas, convierte las filas en diccionario de 
		libro
		"""
		columnasCorrectas = self.validaEstructuraArchivo(hoja)
		arreglo_libros = []
		if columnasCorrectas == True:
			cantidad_filas = hoja.max_row#cuenta la cantidad de filas del excel
			cantidad_columnas = hoja.max_column #cuenta la cantidad de columnas del excel
			for fila in range(2,(cantidad_filas + 1)):
				posicion = 0
				item_libro = {}
				for columna in range(1,(cantidad_columnas + 1)):
					valor = hoja.cell(row = fila, column = columna).value
					if self.columnas[posicion] == 'Cantidad':
						item_libro[self.columnas[posicion]] = int(valor)
					else:
						item_libro[self.columnas[posicion]] = valor
					posicion += 1
				arreglo_libros.append(item_libro)
		else:
			print('Fallo creando el arreglo')
			
		return arreglo_libros



#------------------------------------------------------------------
"""
obj_archivo = Archivo("hoja_de_vida_libros.xlsx")
hoja = obj_archivo.cargarArchivo()
columnasCorrectas = obj_archivo.validaEstructuraArchivo(hoja)
arregloLibros = obj_archivo.convertirExcelEnDiccionario(hoja)
print('')
print(arregloLibros)
"""